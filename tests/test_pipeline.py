"""Kapsam Bot doğrulama testleri.

Çalıştırma:
    python tests/test_pipeline.py
"""

import sqlite3
import sys
import tempfile
import warnings
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE / "app"))
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

import calculations  # noqa: E402
import database  # noqa: E402
import excel_processor as xl  # noqa: E402
import filters  # noqa: E402
from config import (  # noqa: E402
    COL_AC,
    COL_AE,
    COL_AF,
    COL_AG,
    DATA_START_ROW,
    DB_COLUMNS,
    LAST_COL,
    OUTPUT_ROWS,
    OUTPUT_SHEET,
    SOURCE_SHEET,
    SOURCE_XLSX,
    TARGET_SHEET,
    TARGET_XLSX,
)

PASS, FAIL = [], []


def check(name, condition, detail=""):
    if condition:
        PASS.append(name)
        print(f"  [OK]   {name}")
    else:
        FAIL.append((name, detail))
        print(f"  [FAIL] {name}  {detail}")


def make_row(**kw):
    """Test satırı üretir (AH'ye kadar 34 sütun)."""
    row = [None] * LAST_COL
    row[0] = kw.get("id", "TEST-1")
    row[4] = kw.get("e", "")            # E  disiplin
    row[15] = kw.get("p", "")           # P  rapor tipi
    row[16] = kw.get("q", "")           # Q  sahadan gelen bilgi
    row[COL_AC - 1] = kw.get("ac", 0)
    row[COL_AE - 1] = kw.get("ae", 0)
    row[COL_AF - 1] = kw.get("af", 0)
    row[COL_AG - 1] = kw.get("ag", 0)
    return row


# --- 1. Normalizasyon -------------------------------------------------------
print("\n1. Türkçe normalizasyon ve sayı dönüşümü")
check("İ/ı normalize", filters.normalize("İLAVE İŞ") == "ilave is")
check("ş/ğ/ü/ö/ç normalize", filters.normalize("Şğüöç") == "sguoc")
check("satır sonu temizlenir", filters.normalize("Programlı \nİş") == "programli is")
check("None -> boş", filters.normalize(None) == "")
check("contains TR uyumlu", filters.contains("Programlı İş", "programli is"))
check("startswith HD", filters.startswith("HD DEVAM (Haftaya Devam eden uzun iş)", "HD"))
check("startswith Alt", filters.startswith("Alt İŞ ADIMI TAMAM", "Alt"))
check("startswith Tamam", filters.startswith("TAMAM (Ana İş Tamam )", "Tamam"))
check("Tamam, Alt'a eşleşmez", not filters.startswith("TAMAM (Ana İş Tamam )", "Alt"))
check("boş -> 0", filters.to_number(None) == 0.0 and filters.to_number("") == 0.0)
check("metin -> 0", filters.to_number("abc") == 0.0 and filters.to_number("-") == 0.0)
check("virgüllü ondalık", filters.to_number("1.234,5") == 1234.5)
check("#REF! -> 0", filters.to_number("#REF!") == 0.0)

# --- 2. Grup filtreleri -----------------------------------------------------
print("\n2. Grup filtreleri (K / A / L / P / K-1)")
r_k_hd = make_row(id="1", p="Programlı İş", q="HD DEVAM (Haftaya Devam eden uzun iş)", e="MET-DIS", af=500, ag=100)
r_k_alt = make_row(id="2", p="Programlı İş", q="Alt İŞ ADIMI TAMAM", e="TUM-MTL", af=200, ag=150)
r_k_tam = make_row(id="3", p="Programlı İş", q="TAMAM (Ana İş Tamam )", e="INS-DIS", af=300, ag=290)
r_p_only = make_row(id="4", p="Programlı İş", q="DÖY (Daha önce yapıldı)", e="MET-DIS")
r_acil = make_row(id="5", p="Acil İş", e="MET-DIS", af=90)
r_durus = make_row(id="6", p="Duruş İşi", e="ISK-DIS", af=45)
r_ilave = make_row(id="7", p="İlave İş-2 (Saha)", e="IZO-DIS", af=70)
r_steam = make_row(id="8", p="Steam İşleri", e="MET-DIS")
rows = [r_k_hd, r_k_alt, r_k_tam, r_p_only, r_acil, r_durus, r_ilave, r_steam]

check("K: HD ile başlayan", filters.is_group_k(r_k_hd))
check("K: Alt ile başlayan", filters.is_group_k(r_k_alt))
check("K: Tamam ile başlayan", filters.is_group_k(r_k_tam))
check("K: DÖY dahil değil", not filters.is_group_k(r_p_only))
check("K, P'nin alt kümesi", all(filters.is_group_p(r) for r in rows if filters.is_group_k(r)))
check("A: Acil İş", filters.is_group_a(r_acil))
check("A: Duruş İşi", filters.is_group_a(r_durus))
check("A: Programlı İş değil", not filters.is_group_a(r_k_hd))
check("L: İlave iş-2 (Saha)", filters.is_group_l(r_ilave))
check("L: Steam değil", not filters.is_group_l(r_steam))
check("P: Programlı İş", filters.is_group_p(r_p_only))
check("P: Steam dahil değil", not filters.is_group_p(r_steam))
check("K-1: AF-AG=400 > 100", filters.is_group_k1(r_k_hd))
check("K-1: AF-AG=50 dahil değil", not filters.is_group_k1(r_k_alt))
check("K-1: AF-AG=10 dahil değil", not filters.is_group_k1(r_k_tam))
check("K-1 sadece K'dan gelir", not filters.is_group_k1(r_acil))

# --- 3. Alt kümeler ve TUM/DIS ---------------------------------------------
print("\n3. İşyeri kodu alt kümeleri ve TUM/DIS ayrımı")
g = filters.build_groups(rows)
check("K grubu 3 satır", len(g["K"]["_all"]) == 3, f"{len(g['K']['_all'])}")
check("A grubu 2 satır", len(g["A"]["_all"]) == 2, f"{len(g['A']['_all'])}")
check("L grubu 1 satır", len(g["L"]["_all"]) == 1)
check("P grubu 4 satır", len(g["P"]["_all"]) == 4, f"{len(g['P']['_all'])}")
check("K-1 grubu 1 satır", len(g["K1"]["_all"]) == 1)
check("K-metal (MET-DIS) 1", len(g["K"]["MET-DIS"]) == 1)
check("K-tummetal (TUM-MTL) 1", len(g["K"]["TUM-MTL"]) == 1)
check("K-insaat (INS-DIS) 1", len(g["K"]["INS-DIS"]) == 1)
check("A-iskele (ISK-DIS) 1", len(g["A"]["ISK-DIS"]) == 1)
check("L-izole (IZO-DIS) 1", len(g["L"]["IZO-DIS"]) == 1)
check("işyeri kodu tespiti", filters.isyeri_kodu(r_k_alt) == "TUM-MTL")
check("tanımsız kod -> None", filters.isyeri_kodu(make_row(e="KAY-DIS")) is None)

# --- 4. Hesaplamalar --------------------------------------------------------
print("\n4. Toplamlar ve HPU formülleri")
calc_rows = [
    make_row(id="c1", p="Programlı İş", q="TAMAM", e="MET-DIS", ac=1000, ae=100, af=800, ag=600),
    make_row(id="c2", p="Programlı İş", q="HD DEVAM", e="MET-DIS", ac=500, ae=50, af=400, ag=250),
    make_row(id="c3", p="Acil İş", e="MET-DIS", ae=20, af=200),
    make_row(id="c4", p="İlave İş-2 (Saha)", e="MET-DIS", ae=10, af=150),
]
cg = filters.build_groups(calc_rows)
m = calculations.compute_metrics(cg)["MET-DIS"]
check("program_isgucu (AC, P grubu) = 1500", m["program_isgucu"] == 1500, f"{m['program_isgucu']}")
check("kazanilan_sure (AG, K grubu) = 850", m["kazanilan_sure"] == 850, f"{m['kazanilan_sure']}")
check("acil_harcanan (AF, A grubu) = 200", m["acil_harcanan"] == 200, f"{m['acil_harcanan']}")
check("ilave_harcanan (AF, L grubu) = 150", m["ilave_harcanan"] == 150, f"{m['ilave_harcanan']}")
check("fazla_mesai (AE, tüm satırlar) = 180", m["fazla_mesai"] == 180, f"{m['fazla_mesai']}")
# K-1: c1 -> 800-600=200>100 dahil, c2 -> 400-250=150>100 dahil => 350
check("kapsam_artisi (K-1 AF-AG) = 350", m["kapsam_artisi"] == 350, f"{m['kapsam_artisi']}")

out = calculations.compute_output({"MET-DIS": m})
d27 = out["dis"]["D"]
payda = 1500 - 200 - 150 + 180  # 1330
check("HPU = K/(P-A-L+F)", abs(d27["hpu"] - 850 / payda) < 1e-9, f"{d27['hpu']}")
check("HPU Kapsam = (K+K1)/(P-A-L+F)", abs(d27["hpu_kapsam"] - (850 + 350) / payda) < 1e-9)
check("TUM kategorisi MET-DIS almaz", out["tum"]["D"]["program_isgucu"] == 0)
check("I sütunu = D..H toplamı", out["dis"]["I"]["program_isgucu"] == 1500)

zero = calculations.compute_output({"MET-DIS": {k: 0.0 for k in m}})
check("sıfıra bölme çökmez", zero["dis"]["D"]["hpu"] == 0.0 and zero["dis"]["D"]["hpu_kapsam"] == 0.0)

# --- 5. SQLite --------------------------------------------------------------
print("\n5. SQLite şeması ve upsert")
with tempfile.TemporaryDirectory() as tmp:
    db_file = Path(tmp) / "t.db"
    conn = sqlite3.connect(db_file)
    orig = database.DB_PATH
    try:
        cols = ",\n".join(
            f"{n} {t}" + (" PRIMARY KEY" if n == "id_1" else "") for _, n, t in DB_COLUMNS
        )
        conn.execute(f"CREATE TABLE geri_bildirim (\n{cols},\nupdated_at TEXT)")
        conn.commit()

        r1 = make_row(id="X-1", p="Programlı İş", e="MET-DIS", ac=100)
        ins, upd, skip = database.upsert_rows(conn, [r1])
        check("ilk çalıştırma INSERT", (ins, upd) == (1, 0), f"{ins},{upd}")

        r1b = make_row(id="X-1", p="Acil İş", e="MET-DIS", ac=999)
        ins2, upd2, _ = database.upsert_rows(conn, [r1b])
        check("aynı ID-1 tekrar -> UPDATE", (ins2, upd2) == (0, 1), f"{ins2},{upd2}")
        check("duplicate oluşmadı", database.row_count(conn) == 1)
        val = conn.execute("SELECT planlanan_sure_dk, rapor_tipi FROM geri_bildirim").fetchone()
        check("son veri üzerine yazıldı", val[0] == 999.0 and val[1] == "Acil İş", str(val))

        _, _, skipped = database.upsert_rows(conn, [make_row(id="")])
        check("boş ID-1 atlanır", skipped == 1)
        check("program_haftasi sütunu var", "program_haftasi" in [c[1] for c in DB_COLUMNS])
        check("sütun sırası korunur", [c[0] for c in DB_COLUMNS] == list(range(1, 35)))
        check("SQL güvenli isimler", all(
            n.replace("_", "").isalnum() and n.islower() for _, n, _ in DB_COLUMNS
        ))
    finally:
        conn.close()
        database.DB_PATH = orig

# --- 6. Gerçek Excel dosyaları ---------------------------------------------
print("\n6. Gerçek Excel dosyaları")
if SOURCE_XLSX.exists() and TARGET_XLSX.exists():
    import openpyxl

    check("kaynak dosya mevcut", SOURCE_XLSX.exists())
    check("hedef dosya mevcut", TARGET_XLSX.exists())

    wbs = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    check(f"kaynak '{SOURCE_SHEET}' sayfası var", SOURCE_SHEET in wbs.sheetnames)
    wbs.close()

    wbt = openpyxl.load_workbook(TARGET_XLSX, read_only=True, data_only=True)
    check(f"hedef '{TARGET_SHEET}' sayfası var", TARGET_SHEET in wbt.sheetnames)
    check(f"hedef '{OUTPUT_SHEET}' sayfası var", OUTPUT_SHEET in wbt.sheetnames)

    tgt = wbt[TARGET_SHEET]
    first = [c.value for c in next(tgt.iter_rows(min_row=DATA_START_ROW, max_row=DATA_START_ROW, max_col=LAST_COL))]
    check("veri 5. satırdan başlıyor", first[0] not in (None, ""), repr(first[0]))
    check("AH sütununa kadar (34) veri var", len(first) == LAST_COL)
    check("values-only (formül yok)", not any(isinstance(v, str) and v.startswith("=") for v in first))

    ows = wbt[OUTPUT_SHEET]
    written_ok = True
    for metrik, satirlar in OUTPUT_ROWS.items():
        for kategori, row_no in satirlar.items():
            v = ows.cell(row=row_no, column=4).value  # D sütunu
            if not isinstance(v, (int, float)):
                written_ok = False
    check("Output hücrelerinde sayısal değer var", written_ok)
    check("TUM tablosu başlığı satır 5", "KOD" in str(ows.cell(row=5, column=3).value or "").upper())
    check("DIS tablosu başlığı satır 24", "KOD" in str(ows.cell(row=24, column=3).value or "").upper())
    wbt.close()

    db = BASE / "data" / "kapsam_bot.db"
    if db.exists():
        c = sqlite3.connect(db)
        n = c.execute("SELECT COUNT(*) FROM geri_bildirim").fetchone()[0]
        d = c.execute("SELECT COUNT(*) - COUNT(DISTINCT id_1) FROM geri_bildirim").fetchone()[0]
        c.close()
        check("SQLite oluşturuldu ve dolu", n > 0, f"{n} kayıt")
        check("ID-1 duplicate yok", d == 0)
    else:
        print("  [SKIP] SQLite henüz oluşmamış (önce python app/main.py çalıştırın)")
else:
    print("  [SKIP] docs/ içinde Excel dosyaları yok")

# --- Özet -------------------------------------------------------------------
print("\n" + "=" * 62)
print(f"SONUÇ: {len(PASS)} başarılı, {len(FAIL)} başarısız")
print("=" * 62)
for name, detail in FAIL:
    print(f"  BAŞARISIZ: {name}  {detail}")
sys.exit(1 if FAIL else 0)
