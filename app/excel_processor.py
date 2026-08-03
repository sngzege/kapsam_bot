"""Excel okuma/yazma katmanı (openpyxl, values-only aktarım)."""

import shutil
from datetime import datetime

import openpyxl

from config import (
    BACKUP_DIR,
    DATA_START_ROW,
    LAST_COL,
    OUTPUT_ROWS,
    OUTPUT_SHEET,
    SOURCE_SHEET,
    SOURCE_XLSX,
    TARGET_SHEET,
    TARGET_XLSX,
)


class ExcelError(Exception):
    """Kullanıcıya gösterilecek anlaşılır Excel hatası."""


def _require_file(path, rol):
    if not path.exists():
        raise ExcelError(
            f"Gerekli Excel dosyası bulunamadı:\n\n"
            f"  docs/{path.name}   ({rol})\n\n"
            f"Lütfen Excel dosyasını docs klasörüne yerleştirin ve programı tekrar çalıştırın."
        )


def _require_sheet(wb, name, dosya):
    if name not in wb.sheetnames:
        raise ExcelError(
            f"'{dosya}' dosyasında '{name}' worksheet'i bulunamadı.\n"
            f"Mevcut sayfalar: {', '.join(wb.sheetnames)}"
        )


def read_source():
    """Kaynak Excel'i values-only okur.

    Dönüş: 5. satırdan itibaren, AH sütununa kadar dolu satırların listesi.
    """
    _require_file(SOURCE_XLSX, "veri kaynağı")
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    try:
        _require_sheet(wb, SOURCE_SHEET, SOURCE_XLSX.name)
        ws = wb[SOURCE_SHEET]
        rows = []
        for row in ws.iter_rows(
            min_row=DATA_START_ROW, max_col=LAST_COL, values_only=True
        ):
            # ID-1 (A sütunu) boşsa satır veri değildir
            if row[0] is None or str(row[0]).strip() == "":
                continue
            rows.append(list(row))
        return rows
    finally:
        wb.close()


def backup_target():
    """HPU.xlsx dosyasının zaman damgalı yedeğini alır."""
    _require_file(TARGET_XLSX, "hedef dosya")
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = BACKUP_DIR / f"HPU_{stamp}.xlsx"
    shutil.copy2(TARGET_XLSX, dest)
    return dest


def open_target():
    """Hedef workbook'u formülleri koruyarak açar."""
    _require_file(TARGET_XLSX, "hedef dosya")
    try:
        wb = openpyxl.load_workbook(TARGET_XLSX, data_only=False)
    except Exception as exc:  # bozuk/kilitli dosya
        raise ExcelError(
            f"'{TARGET_XLSX.name}' açılamadı: {exc}\n"
            f"Dosya Excel/LibreOffice'te açıksa kapatıp tekrar deneyin."
        ) from exc
    _require_sheet(wb, TARGET_SHEET, TARGET_XLSX.name)
    _require_sheet(wb, OUTPUT_SHEET, TARGET_XLSX.name)
    return wb


def write_data_sheet(wb, rows):
    """Veriyi hedef 'Geri Bildirim' sayfasına 5. satırdan itibaren yazar.

    Values-only aktarım. Eski veri temizlenerek duplicate oluşması engellenir.
    """
    ws = wb[TARGET_SHEET]

    # Eski veri bloğunu temizle (5. satırdan mevcut son satıra kadar)
    old_last = ws.max_row
    for r in range(DATA_START_ROW, old_last + 1):
        for c in range(1, LAST_COL + 1):
            ws.cell(row=r, column=c).value = None

    # Yeni veriyi yaz
    for i, row in enumerate(rows):
        target_row = DATA_START_ROW + i
        for c in range(1, LAST_COL + 1):
            value = row[c - 1] if c - 1 < len(row) else None
            ws.cell(row=target_row, column=c).value = value

    return len(rows)


def verify_output_template(wb):
    """Output şablonunun beklenen yapıda olduğunu doğrular.

    Şablon ile dokümandaki satır tanımları çelişirse uyarı listesi döner.
    """
    ws = wb[OUTPUT_SHEET]
    warnings = []
    # TUM tablosu 5. satırda, DIS tablosu 24. satırda başlık taşımalı
    for header_row, kategori in ((5, "TUM"), (24, "DIS")):
        label = ws.cell(row=header_row, column=3).value  # C sütunu = "KOD"
        if label is None or "KOD" not in str(label).upper():
            warnings.append(
                f"Output şablonu beklenenden farklı: C{header_row} 'KOD' değil ({label!r})"
            )
    return warnings


def write_output(wb, output):
    """Hesaplanan değerleri Output sayfasındaki hücrelere yazar.

    Dönüş: yazılan hücre sayısı.
    """
    ws = wb[OUTPUT_SHEET]
    written = 0
    for metrik, satirlar in OUTPUT_ROWS.items():
        for kategori, row_no in satirlar.items():
            cols = output.get(kategori, {})
            for col_letter, values in cols.items():
                if metrik not in values:
                    continue
                ws[f"{col_letter}{row_no}"] = round(values[metrik], 6)
                written += 1
    return written


def save_target(wb):
    """Hedef workbook'u diske yazar."""
    try:
        wb.save(TARGET_XLSX)
    except PermissionError as exc:
        raise ExcelError(
            f"'{TARGET_XLSX.name}' kaydedilemedi. Dosya başka bir programda açık olabilir.\n"
            f"Excel/LibreOffice'te kapatıp tekrar çalıştırın."
        ) from exc
    finally:
        wb.close()
